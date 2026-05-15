from fastapi import APIRouter, Depends, HTTPException, status, Body
from fastapi.security import OAuth2PasswordBearer
from sqlalchemy.orm import Session
from sqlalchemy import and_
from auth_utils import hash_password, verify_password, create_access_token, verify_access_token
from config import get_db
from schemas.models import *
from datetime import timedelta
from typing import List, Optional, Tuple
import re
from routes.auth import oauth2_scheme

router = APIRouter()

@router.get("/", response_model=List[dict])
async def get_all_users(
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Get all users"""
    user_data = verify_access_token(token)
    if not user_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    # Check if the user is an admin
    if user_data.get("type") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view all users")
    
    users = db.query(UserInDB).all()
    
    result = []
    for user in users:
        # Не включаем хэш пароля в ответ
        result.append({
            "id": user.id,
            "name": user.name,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "type": user.type,
            "company_name": user.company_name,
            "shanyrak": user.shanyrak,
            "is_active": user.is_active,
            "created_at": user.created_at,
            "updated_at": user.updated_at
        })
    
    return result

@router.get("/by-type/{user_type}", response_model=List[dict])
async def get_users_by_type(
    user_type: str,
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Get users by type (curator, teacher, admin, etc.)"""
    user_data = verify_access_token(token)
    if not user_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    if user_data.get("type") not in ("admin", "curator"):
        raise HTTPException(status_code=403, detail="Only admins and curators can list users by type")

    # Validate user type
    valid_types = ['admin', 'curator', 'teacher', 'user']
    if user_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid user type. Must be one of: {valid_types}")
    
    users = db.query(UserInDB).filter(
        UserInDB.type == user_type,
        UserInDB.is_active == 1
    ).all()
    
    result = []
    for user in users:
        # Get additional info based on user type
        additional_info = {}
        
        if user_type == 'curator':
            # Count assigned grades
            grade_count = db.query(GradeInDB).filter(GradeInDB.curator_id == user.id).count()
            additional_info['assigned_grades_count'] = grade_count
        
        elif user_type == 'teacher':
            # Get teaching assignments with subject and grade details
            assignments = db.query(TeacherAssignmentInDB).filter(
                TeacherAssignmentInDB.teacher_id == user.id,
                TeacherAssignmentInDB.is_active == 1
            ).all()
            additional_info['assignment_count'] = len(assignments)

            subjects_set = set()
            grades_set = set()
            for a in assignments:
                if a.subject_id:
                    subj = db.query(SubjectInDB).filter(SubjectInDB.id == a.subject_id).first()
                    if subj:
                        subjects_set.add(subj.name)
                if a.grade_id:
                    grade = db.query(GradeInDB).filter(GradeInDB.id == a.grade_id).first()
                    if grade:
                        grades_set.add(f"{grade.grade}{grade.parallel}")
            additional_info['subjects'] = sorted(subjects_set)
            additional_info['grades'] = sorted(grades_set)
        
        result.append({
            "id": user.id,
            "name": user.name,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "type": user.type,
            "company_name": user.company_name,
            "shanyrak": user.shanyrak,
            "is_active": user.is_active,
            "created_at": user.created_at,
            "updated_at": user.updated_at,
            **additional_info
        })
    
    return result

@router.get("/{user_id}", response_model=dict)
async def get_user(
    user_id: int,
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Get user by ID"""
    user_data = verify_access_token(token)
    if not user_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    # Admins can view any user, regular users can only view themselves
    if user_data.get("type") != "admin" and user_data.get("id") != user_id:
        raise HTTPException(status_code=403, detail="You can only view your own data")
    
    user = db.query(UserInDB).filter(UserInDB.id == user_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "id": user.id,
        "name": user.name,
        "email": user.email,
        "type": user.type
    }

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    type: str

@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_user(
    user_data: UserCreate,
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Create a new user"""
    admin_data = verify_access_token(token)
    if not admin_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    # Only admins can create users
    if admin_data.get("type") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can create users")
    
    # Check if email already exists
    existing_user = db.query(UserInDB).filter(UserInDB.email == user_data.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create new user
    hashed_password = hash_password(user_data.password)
    
    new_user = UserInDB(
        name=user_data.name,
        email=user_data.email,
        hashed_password=hashed_password,
        type=user_data.type
    )
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return {
        "id": new_user.id,
        "message": "User created successfully"
    }

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    type: Optional[str] = None

@router.put("/{user_id}", status_code=status.HTTP_200_OK)
async def update_user(
    user_id: int,
    update_data: UserUpdate,
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Update user by ID"""
    user_data = verify_access_token(token)
    if not user_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    # Admins can update any user, regular users can only update themselves and not change their type
    if user_data.get("type") != "admin":
        if user_data.get("id") != user_id:
            raise HTTPException(status_code=403, detail="You can only update your own data")
        if update_data.type is not None:
            raise HTTPException(status_code=403, detail="You cannot change your user type")
    
    user = db.query(UserInDB).filter(UserInDB.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update email if provided and not already taken
    if update_data.email is not None and update_data.email != user.email:
        existing_user = db.query(UserInDB).filter(UserInDB.email == update_data.email).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Email already registered")
        user.email = update_data.email
    
    # Update name if provided
    if update_data.name is not None:
        user.name = update_data.name
    
    # Update type if provided (admin only)
    if update_data.type is not None and user_data.get("type") == "admin":
        user.type = update_data.type
    
    # Update password if provided
    if update_data.password is not None:
        user.hashed_password = hash_password(update_data.password)
    
    db.commit()
    db.refresh(user)
    
    return {
        "message": "User updated successfully"
    }

@router.delete("/{user_id}", status_code=status.HTTP_200_OK)
async def delete_user(
    user_id: int,
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Delete user by ID"""
    admin_data = verify_access_token(token)
    if not admin_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    # Only admins can delete users
    if admin_data.get("type") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    # Prevent admin from deleting themselves
    if admin_data.get("id") == user_id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    
    user = db.query(UserInDB).filter(UserInDB.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.delete(user)
    db.commit()
    
    return {
        "message": "User deleted successfully"
    }

from fastapi import File, UploadFile
import openpyxl
import io

class BulkUploadRowResult(BaseModel):
    row: int
    status: str  # "created" | "updated" | "skipped"
    message: str
    teacher_name: Optional[str] = None
    class_name: Optional[str] = None
    subject_name: Optional[str] = None
    subject_group_name: Optional[str] = None


class BulkUploadResult(BaseModel):
    success: bool
    total_processed: int
    created_count: int
    updated_count: int
    error_count: int
    errors: List[dict]
    created_users: List[dict]
    row_results: List[BulkUploadRowResult] = []


def _parse_class_cell(cell_val) -> Tuple[Optional[str], Optional[str]]:
    """Parse class cell (e.g. '11A', '10B') into (grade, parallel). Returns (None, None) if invalid."""
    if cell_val is None or str(cell_val).strip() == "":
        return None, None
    s = str(cell_val).strip()
    m = re.match(r"^(\d{1,2})\s*([A-Za-zА-Яа-яЁёІіҢңҒғҚқӨөҰұҮүҺһ]?)\s*$", s)
    if m:
        grade = m.group(1)
        parallel = m.group(2) if m.group(2) else "A"
        return grade, parallel
    if re.match(r"^\d{1,2}$", s):
        return s, "A"
    return None, None


def _parse_class_list_cell(cell_val) -> Tuple[List[Tuple[str, str]], List[str]]:
    """
    Parse a class cell that may contain multiple classes separated by commas.
    Example: "11A, 11B, 12A" -> [("11","A"), ("11","B"), ("12","A")]
    Returns (valid_classes, invalid_tokens)
    """
    if cell_val is None or str(cell_val).strip() == "":
        return [], []

    raw_value = str(cell_val).strip()
    tokens = [t.strip() for t in re.split(r'[,;/\n]+', raw_value) if t and t.strip()]
    valid_classes: List[Tuple[str, str]] = []
    invalid_tokens: List[str] = []

    for token in tokens:
        grade_str, parallel_str = _parse_class_cell(token)
        if grade_str and parallel_str:
            valid_classes.append((grade_str, parallel_str))
        else:
            invalid_tokens.append(token)

    return valid_classes, invalid_tokens


def _get_or_create_grade(db: Session, grade_str: str, parallel: str, user_id: int) -> Optional[GradeInDB]:
    existing = db.query(GradeInDB).filter(
        GradeInDB.grade == grade_str,
        GradeInDB.parallel == parallel
    ).first()
    if existing:
        return existing
    new_grade = GradeInDB(
        grade=grade_str,
        parallel=parallel,
        user_id=user_id,
        student_count=0
    )
    db.add(new_grade)
    db.flush()
    db.refresh(new_grade)
    return new_grade


def _get_or_create_subject(db: Session, name: str) -> Optional[SubjectInDB]:
    if not name or not str(name).strip():
        return None
    existing = db.query(SubjectInDB).filter(SubjectInDB.name == name.strip()).first()
    if existing:
        return existing
    new_subject = SubjectInDB(
        name=name.strip(),
        applicable_parallels=[],
        allows_subject_groups=False,
        is_active=1
    )
    db.add(new_subject)
    db.flush()
    db.refresh(new_subject)
    return new_subject


def _get_or_create_subject_group(db: Session, grade_id: int, subject_id: int, name: str) -> Optional[object]:
    from schemas.models import SubjectGroupInDB
    if not name or not str(name).strip():
        return None
    existing = db.query(SubjectGroupInDB).filter(
        SubjectGroupInDB.grade_id == grade_id,
        SubjectGroupInDB.subject_id == subject_id,
        SubjectGroupInDB.name == name.strip(),
        SubjectGroupInDB.is_active == 1
    ).first()
    if existing:
        return existing
    new_group = SubjectGroupInDB(
        grade_id=grade_id,
        subject_id=subject_id,
        name=name.strip(),
        is_active=1
    )
    db.add(new_group)
    db.flush()
    db.refresh(new_group)
    return new_group


@router.post("/bulk-upload-teachers", response_model=BulkUploadResult)
async def bulk_upload_teachers(
    file: UploadFile = File(...),
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """
    Bulk upload teachers from Excel file.
    Expected columns: ФИО | Email | Классы (через запятую) | Предмет
    A teacher with multiple subjects appears on multiple rows (same ФИО/Email, different subject/classes).
    """
    admin_data = verify_access_token(token)
    if not admin_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    if admin_data.get("type") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can bulk upload users")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    creator_user = db.query(UserInDB).filter(UserInDB.email == admin_data.get("sub")).first()
    if not creator_user:
        creator_user = db.query(UserInDB).filter(UserInDB.id == admin_data.get("id")).first()
    creator_id = creator_user.id if creator_user else admin_data.get("id", 1)

    # Load list of valid subjects from the "Предметы" sheet (if present)
    valid_subjects: set = set()

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))

        # Collect valid subjects from the reference sheet
        for sname in ("Предметы", "предметы", "Subjects"):
            if sname in workbook.sheetnames:
                subj_sheet = workbook[sname]
                for row in subj_sheet.iter_rows(min_row=2, values_only=True):
                    val = row[0] if row else None
                    if val and str(val).strip():
                        valid_subjects.add(str(val).strip())
                break

        # Find the teachers data sheet (first sheet, or named "Учителя"/"Teachers")
        sheet = workbook.active
        for sname in ("Учителя", "учителя", "Teachers"):
            if sname in workbook.sheetnames:
                sheet = workbook[sname]
                break

        created_users = []
        errors = []
        row_results = []
        total_processed = 0

        # Detect header row: find row containing "ФИО"
        header_row_idx = 1
        for check_row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and any(cell and "фио" in str(cell).strip().lower() for cell in row):
                header_row_idx = check_row_idx
                break

        # Build column index map from header
        header = [str(c).strip().lower() if c else "" for c in next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]
        col_map = {}
        for idx, h in enumerate(header):
            if "фио" in h or h == "имя" or h == "name":
                col_map["fio"] = idx
            elif "email" in h or "почта" in h or "e-mail" in h:
                col_map["email"] = idx
            elif "класс" in h or "class" in h:
                col_map["classes"] = idx
            elif "предмет" in h or "subject" in h:
                col_map["subject"] = idx

        if "fio" not in col_map or "email" not in col_map:
            raise HTTPException(status_code=400, detail="Не найдены обязательные столбцы: ФИО и Email. Проверьте заголовки.")

        try:
            # ── Overwrite mode: deactivate ALL existing teacher assignments ──
            db.query(TeacherAssignmentInDB).filter(
                TeacherAssignmentInDB.is_active == 1
            ).update({"is_active": 0})
            db.flush()

            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                total_processed += 1

                fio = str(row[col_map["fio"]]).strip() if col_map.get("fio") is not None and len(row) > col_map["fio"] and row[col_map["fio"]] else None
                email = str(row[col_map["email"]]).strip().lower() if col_map.get("email") is not None and len(row) > col_map["email"] and row[col_map["email"]] else None
                class_cell = row[col_map["classes"]] if col_map.get("classes") is not None and len(row) > col_map["classes"] else None
                subject_cell = str(row[col_map["subject"]]).strip() if col_map.get("subject") is not None and len(row) > col_map["subject"] and row[col_map["subject"]] else None

                if not fio or not email:
                    err = {"row": row_idx, "error": "Отсутствуют обязательные поля (ФИО или Email)", "data": {"fio": fio, "email": email}}
                    errors.append(err)
                    row_results.append(BulkUploadRowResult(row=row_idx, status="skipped", message=err["error"]))
                    continue

                # Validate subject against reference sheet if available
                if subject_cell and valid_subjects and subject_cell not in valid_subjects:
                    err = {
                        "row": row_idx,
                        "error": f"Предмет '{subject_cell}' не найден в списке предметов",
                        "data": {"fio": fio, "email": email, "subject": subject_cell}
                    }
                    errors.append(err)
                    row_results.append(BulkUploadRowResult(row=row_idx, status="skipped", message=err["error"]))
                    continue

                class_pairs, invalid_class_tokens = _parse_class_list_cell(class_cell)
                if invalid_class_tokens:
                    err = {
                        "row": row_idx,
                        "error": f"Некорректные классы: {', '.join(invalid_class_tokens)}",
                        "data": {"fio": fio, "email": email, "class_cell": class_cell}
                    }
                    errors.append(err)
                    row_results.append(BulkUploadRowResult(row=row_idx, status="skipped", message=err["error"]))
                    continue

                try:
                    existing_user = db.query(UserInDB).filter(UserInDB.email == email).first()
                    if existing_user:
                        teacher = existing_user
                        if teacher.type not in ("teacher", "admin"):
                            teacher.type = "teacher"
                    else:
                        fio_parts = fio.split()
                        last_name = fio_parts[0] if fio_parts else ""
                        first_name = fio_parts[1] if len(fio_parts) > 1 else ""
                        hashed_password = hash_password("123")
                        teacher = UserInDB(
                            name=fio,
                            first_name=first_name,
                            last_name=last_name,
                            email=email,
                            hashed_password=hashed_password,
                            type="teacher",
                        )
                        db.add(teacher)
                        db.flush()
                        db.refresh(teacher)
                        created_users.append({"id": teacher.id, "name": fio, "email": email, "subject": subject_cell})

                    subject_id = None
                    class_display = ",".join([f"{g}{p}" for g, p in class_pairs]) if class_pairs else None

                    if subject_cell:
                        subj = _get_or_create_subject(db, subject_cell)
                        if subj:
                            subject_id = subj.id

                    if subject_id:
                        if not class_pairs:
                            # Subject-only assignment (no classes specified)
                            filters = [
                                TeacherAssignmentInDB.teacher_id == teacher.id,
                                TeacherAssignmentInDB.subject_id == subject_id,
                                TeacherAssignmentInDB.is_active == 1,
                                TeacherAssignmentInDB.subgroup_id.is_(None),
                                TeacherAssignmentInDB.grade_id.is_(None),
                                TeacherAssignmentInDB.subject_group_id.is_(None),
                            ]
                            existing_assignment = db.query(TeacherAssignmentInDB).filter(and_(*filters)).first()
                            if not existing_assignment:
                                new_assignment = TeacherAssignmentInDB(
                                    teacher_id=teacher.id,
                                    subject_id=subject_id,
                                    grade_id=None,
                                    subject_group_id=None,
                                    subgroup_id=None,
                                    is_active=1,
                                )
                                db.add(new_assignment)
                        else:
                            for grade_str, parallel_str in class_pairs:
                                db_grade = _get_or_create_grade(db, grade_str, parallel_str, creator_id)
                                if not db_grade:
                                    continue

                                filters = [
                                    TeacherAssignmentInDB.teacher_id == teacher.id,
                                    TeacherAssignmentInDB.subject_id == subject_id,
                                    TeacherAssignmentInDB.is_active == 1,
                                    TeacherAssignmentInDB.subgroup_id.is_(None),
                                    TeacherAssignmentInDB.grade_id == db_grade.id,
                                    TeacherAssignmentInDB.subject_group_id.is_(None),
                                ]
                                existing_assignment = db.query(TeacherAssignmentInDB).filter(and_(*filters)).first()
                                if not existing_assignment:
                                    new_assignment = TeacherAssignmentInDB(
                                        teacher_id=teacher.id,
                                        subject_id=subject_id,
                                        grade_id=db_grade.id,
                                        subject_group_id=None,
                                        subgroup_id=None,
                                        is_active=1,
                                    )
                                    db.add(new_assignment)

                    status_msg = "updated" if existing_user else "created"
                    row_results.append(BulkUploadRowResult(
                        row=row_idx,
                        status=status_msg,
                        message="OK",
                        teacher_name=fio,
                        class_name=class_display,
                        subject_name=subject_cell,
                    ))

                except Exception as e:
                    err = {"row": row_idx, "error": str(e), "data": {"fio": fio, "email": email}}
                    errors.append(err)
                    row_results.append(BulkUploadRowResult(row=row_idx, status="skipped", message=str(e)))

            # Single commit at the end — all or nothing
            db.commit()

        except HTTPException:
            db.rollback()
            raise
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Ошибка при обработке данных, все изменения отменены: {str(e)}")

        updated_count = sum(1 for r in row_results if r.status == "updated")
        return BulkUploadResult(
            success=len(errors) == 0,
            total_processed=total_processed,
            created_count=len(created_users),
            updated_count=updated_count,
            error_count=len(errors),
            errors=errors,
            created_users=created_users,
            row_results=row_results,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process Excel file: {str(e)}")


from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


@router.get("/teachers-template")
async def download_teachers_template(
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db)
):
    """Download Excel template for bulk teacher upload. Includes a 'Предметы' reference sheet."""
    admin_data = verify_access_token(token)
    if not admin_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    if admin_data.get("type") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can download templates")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Учителя (teachers data) ──
    ws = wb.active
    ws.title = "Учителя"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["ФИО", "Email", "Классы", "Предмет"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Example data showing a teacher with multiple subjects on separate rows
    examples = [
        ("Иванов Иван Иванович", "ivanov_i@school.edu.kz", "9A,10A,10B", "Математика"),
        ("Иванов Иван Иванович", "ivanov_i@school.edu.kz", "11A,11B", "Физика"),
        ("Петрова Анна Сергеевна", "petrova_a@school.edu.kz", "7A,7B,8A", "Английский язык"),
        ("Сидоров Алексей Петрович", "sidorov_a@school.edu.kz", "10A,10B,11A,12A", "Информатика"),
    ]

    for row_idx, (fio, email, classes, subject) in enumerate(examples, 2):
        ws.cell(row=row_idx, column=1, value=fio).border = thin_border
        ws.cell(row=row_idx, column=2, value=email).border = thin_border
        ws.cell(row=row_idx, column=3, value=classes).border = thin_border
        ws.cell(row=row_idx, column=4, value=subject).border = thin_border

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25

    # Add instructions row
    note_row = len(examples) + 3
    note_cell = ws.cell(row=note_row, column=1, value="Инструкция:")
    note_cell.font = Font(bold=True, color="FF0000")
    ws.cell(row=note_row + 1, column=1, value="1. Если у учителя несколько предметов — заполните отдельную строку для каждого предмета (ФИО и Email одинаковые).")
    ws.cell(row=note_row + 2, column=1, value="2. Классы указывайте через запятую в одной ячейке (например: 9A,10B,11C).")
    ws.cell(row=note_row + 3, column=1, value="3. Предметы берутся из листа 'Предметы'. Убедитесь, что название предмета совпадает.")
    ws.cell(row=note_row + 4, column=1, value="4. Пароль по умолчанию для новых учителей: 123")

    # ── Sheet 2: Предметы (subjects reference) ──
    ws_subj = wb.create_sheet("Предметы")

    subj_header_cell = ws_subj.cell(row=1, column=1, value="Предмет")
    subj_header_cell.font = header_font
    subj_header_cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    subj_header_cell.alignment = Alignment(horizontal='center')
    subj_header_cell.border = thin_border

    # Load existing subjects from DB
    subjects = db.query(SubjectInDB).filter(SubjectInDB.is_active == 1).order_by(SubjectInDB.name).all()
    if subjects:
        for s_idx, subj in enumerate(subjects, 2):
            cell = ws_subj.cell(row=s_idx, column=1, value=subj.name)
            cell.border = thin_border
    else:
        # Provide default subjects if none exist yet
        defaults = [
            "Математика", "Физика", "Химия", "Биология", "Информатика",
            "Английский язык", "Казахский язык", "Русский язык", "История",
            "География", "Физическая культура", "Искусство", "НВП",
        ]
        for s_idx, name in enumerate(defaults, 2):
            cell = ws_subj.cell(row=s_idx, column=1, value=name)
            cell.border = thin_border

    ws_subj.column_dimensions['A'].width = 30

    # ── Add dropdown (data validation) for "Предмет" column on the teachers sheet ──
    subj_count = len(subjects) if subjects else len(defaults)
    subj_last_row = subj_count + 1  # +1 because data starts at row 2
    dv = DataValidation(
        type="list",
        formula1=f"'Предметы'!$A$2:$A${subj_last_row}",
        allow_blank=True
    )
    dv.error = "Выберите предмет из списка (лист 'Предметы')"
    dv.errorTitle = "Некорректный предмет"
    dv.prompt = "Выберите предмет из списка"
    dv.promptTitle = "Предмет"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"D2:D500")

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teachers_template.xlsx"}
    )


