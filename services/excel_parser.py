import pandas as pd
from fastapi import HTTPException
from io import BytesIO
import json
import re
from typing import Dict, List, Optional, Tuple, Any, Union
import unicodedata

def normalize_name(name: str) -> str:
    """Normalize student name for consistent storage"""
    if not name or pd.isna(name):
        return ""
    
    # Convert to string and strip whitespace
    name = str(name).strip()
    
    # Unicode normalization
    name = unicodedata.normalize('NFKC', name)
    
    # Remove extra spaces between words
    name = re.sub(r'\s+', ' ', name)
    
    # Remove any trailing commas or periods
    name = name.rstrip('.,')
    
    return name

def validate_percentage(value: Any) -> Optional[float]:
    """Validate and convert percentage value"""
    if pd.isna(value) or value == "" or value is None:
        return None
    
    try:
        # Convert to float
        percent = float(value)
        
        # Validate range (0-100)
        if 0 <= percent <= 100:
            return percent
        else:
            return None
    except (ValueError, TypeError):
        return None

def calculate_predicted_scores_by_quarter(
    previous_class_score: Optional[float],
    current_quarters: List[Optional[float]],
    teacher_percent: Optional[float],
    weights: Dict[str, float] = None
) -> List[float]:
    """
    Calculate predicted scores for each quarter based on available data
    
    Formula for each quarter Qn:
    P(Qn) = (w_prev * prev_class_score) + (w_teacher * teacher_score) + (w_quarters * avg(Q1...Qn-1))
    
    Args:
        previous_class_score: Score from previous class (previous year)
        current_quarters: List of 4 quarter scores (can contain None for future quarters)
        teacher_percent: Teacher's assessment percentage
        weights: Dictionary with weights for 'previous_class', 'teacher', 'quarters'
    
    Returns:
        List of 4 predicted scores, one for each quarter
    """
    if weights is None:
        weights = {
            'previous_class': 0.3,
            'teacher': 0.2,
            'quarters': 0.5
        }
    
    w_prev = weights.get('previous_class', 0.3)
    w_teacher = weights.get('teacher', 0.2)
    w_quarters = weights.get('quarters', 0.5)

    predicted_scores = [0.0] * 4

    # --- P(Q1) ---
    # For Q1, prediction is based on previous class and teacher scores.
    # We need to normalize the weights as there are no previous quarters.
    base_prediction_components = []
    total_base_weight = 0
    
    if previous_class_score is not None:
        base_prediction_components.append(w_prev * previous_class_score)
        total_base_weight += w_prev
        
    if teacher_percent is not None:
        base_prediction_components.append(w_teacher * teacher_percent)
        total_base_weight += w_teacher

    if total_base_weight > 0:
        normalized_prediction = sum(base_prediction_components) / total_base_weight
        predicted_scores[0] = round(normalized_prediction, 1)
    else:
        predicted_scores[0] = 0.0

    # --- P(Q2), P(Q3), P(Q4) ---
    for i in range(1, 4):
        prev_quarters = [q for q in current_quarters[:i] if q is not None and q > 0]
        
        if not prev_quarters:
            # If no previous quarters, prediction is the same as for Q1
            predicted_scores[i] = predicted_scores[0]
            continue
            
        avg_prev_quarters = sum(prev_quarters) / len(prev_quarters)
        
        prediction = (
            (w_prev * previous_class_score if previous_class_score is not None else 0) +
            (w_teacher * teacher_percent if teacher_percent is not None else 0) +
            (w_quarters * avg_prev_quarters)
        )
        predicted_scores[i] = round(prediction, 1)

    return predicted_scores


def load_prediction_weights_from_db(db: Any) -> Dict[str, float]:
    """Веса из настроек (как при импорте Excel)."""
    from schemas.models import PredictionSettings

    prediction_settings = db.query(PredictionSettings).filter(
        PredictionSettings.is_active == 1
    ).first()
    weights: Dict[str, float] = {
        "previous_class": 0.3,
        "teacher": 0.2,
        "quarters": 0.5,
    }
    if prediction_settings and prediction_settings.weights:
        weights = prediction_settings.weights
    return weights


def recalculate_predicted_and_danger_from_actual(
    actual_scores_list: Union[List[Any], None],
    previous_class_score: Optional[float],
    teacher_percent: Optional[float],
    weights: Optional[Dict[str, float]] = None,
) -> Tuple[List[float], int, float]:
    """
    Та же логика, что при импорте Excel: прогноз по четвертям + уровень риска по заполненным четвертям.
    Для прогноза пустые ячейки — None (0 или отсутствие значения не считаются введённой оценкой).
    """
    raw_in: List[Any] = list(actual_scores_list or [])[:4]
    while len(raw_in) < 4:
        raw_in.append(0.0)

    quarters_opt: List[Optional[float]] = []
    for x in raw_in:
        if x is None:
            quarters_opt.append(None)
        else:
            try:
                fv = float(x)
            except (TypeError, ValueError):
                quarters_opt.append(None)
                continue
            quarters_opt.append(fv if fv > 0 else None)

    w = weights or {
        "previous_class": 0.3,
        "teacher": 0.2,
        "quarters": 0.5,
    }
    predicted = calculate_predicted_scores_by_quarter(
        previous_class_score,
        quarters_opt,
        teacher_percent,
        w,
    )

    raw = []
    for x in raw_in:
        if x is None:
            raw.append(0.0)
        else:
            try:
                raw.append(float(x))
            except (TypeError, ValueError):
                raw.append(0.0)

    actual_completed = [s for s in raw if s > 0]
    num_completed = len(actual_completed)
    danger_level = 0
    percentage_difference = 0.0

    if num_completed > 0 and len(predicted) >= num_completed:
        predicted_for_completed = predicted[:num_completed]
        avg_actual = sum(actual_completed) / num_completed
        avg_predicted = sum(predicted_for_completed) / num_completed
        delta = avg_actual - avg_predicted
        if delta < -15:
            danger_level = 3
        elif delta < -10:
            danger_level = 2
        elif delta < -5:
            danger_level = 1
        else:
            danger_level = 0
        if avg_predicted > 0:
            percentage_difference = (delta / avg_predicted) * 100

    return predicted, danger_level, round(percentage_difference, 1)


def parse_excel_grades(
    file_content: bytes,
    expected_columns: Dict[str, List[str]] = None,
    weights: Dict[str, float] = None
) -> Dict[str, Any]:
    """
    Parse Excel file for grade upload
    
    Expected columns (case-insensitive, flexible matching):
    - ФИО / Name / Student Name
    - Процент за 1 предыдущий класс / Previous Class Score / Previous Year %
    - Q1 / Четверть 1 / Quarter 1
    - Q2 / Четверть 2 / Quarter 2  
    - Q3 / Четверть 3 / Quarter 3
    - Q4 / Четверть 4 / Quarter 4
    - Учитель / Teacher / Teacher %
    
    Args:
        file_content: Excel file content as bytes
        expected_columns: Dictionary mapping field names to possible column name aliases
        weights: Dictionary with weights for prediction calculation
    """
    
    if expected_columns is None:
        expected_columns = {
            'name': ['фио', 'имя', 'name', 'student', 'студент', 'ученик'],
            'previous_class': ['процент за 1 предыдущий класс', 'previous class', 'previous year', 'предыдущий класс', 'предыдущий год', 'prev class'],
            'q1': ['q1', 'четверть 1', 'quarter 1', '1 четверть', 'ч1'],
            'q2': ['q2', 'четверть 2', 'quarter 2', '2 четверть', 'ч2'],
            'q3': ['q3', 'четверть 3', 'quarter 3', '3 четверть', 'ч3'],
            'q4': ['q4', 'четверть 4', 'quarter 4', '4 четверть', 'ч4'],
            'teacher': ['учитель', 'teacher', 'преподаватель', 'препод']
        }
    
    if weights is None:
        weights = {
            'previous_class': 0.3,
            'teacher': 0.2,
            'quarters': 0.5
        }
    
    try:
        # Read Excel file
        excel_data = BytesIO(file_content)
        df = pd.read_excel(excel_data, sheet_name=0, header=0)
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty")
        
        # Normalize column names for matching
        df.columns = df.columns.astype(str)
        normalized_columns = {col: col.lower().strip() for col in df.columns}
        
        # Find column mappings
        column_mapping = {}
        
        for field, possible_names in expected_columns.items():
            found = False
            for col_name, normalized in normalized_columns.items():
                if any(possible in normalized for possible in possible_names):
                    column_mapping[field] = col_name
                    found = True
                    break
            
            if not found and field == 'name':
                # Name column is required
                raise HTTPException(
                    status_code=400, 
                    detail=f"Required column not found: {field}. Available columns: {list(df.columns)}"
                )
        
        # Parse student data
        students_data = []
        warnings = []
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if row.isna().all():
                    continue
                
                # Get student name (required)
                student_name = normalize_name(row.get(column_mapping.get('name', '')))
                if not student_name:
                    warnings.append(f"Row {index + 2}: Missing student name, skipped")
                    continue
                
                # Parse percentages
                previous_class_score = validate_percentage(row.get(column_mapping.get('previous_class')))
                teacher_percent = validate_percentage(row.get(column_mapping.get('teacher')))
                
                # Parse quarterly grades
                quarters = []
                for quarter in ['q1', 'q2', 'q3', 'q4']:
                    col = column_mapping.get(quarter)
                    if col:
                        quarter_value = validate_percentage(row.get(col))
                        quarters.append(quarter_value)
                    else:
                        quarters.append(None)
                
                # Calculate predicted scores for each quarter
                predicted_scores = calculate_predicted_scores_by_quarter(
                    previous_class_score, quarters, teacher_percent, weights
                )
                
                # Prepare actual scores (replace None with 0.0 for calculations)
                actual_scores = [q if q is not None else 0.0 for q in quarters]
                
                student_data = {
                    "student_name": student_name,
                    "previous_class_score": previous_class_score,
                    "current_quarters": quarters,
                    "teacher_percent": teacher_percent,
                    "actual_scores": actual_scores,
                    "predicted_scores": predicted_scores
                }
                
                students_data.append(student_data)
                
            except Exception as e:
                errors.append(f"Row {index + 2}: Error processing data - {str(e)}")
                continue
        
        if not students_data:
            raise HTTPException(
                status_code=400, 
                detail="No valid student data found in Excel file"
            )
        
        response = {
            "students": students_data,
            "total_rows": len(df),
            "processed_rows": len(students_data),
            "warnings": warnings,
            "errors": errors,
            "column_mapping": column_mapping
        }
        
        return response
        
    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel file is empty or corrupted")
    except pd.errors.ParserError as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Unexpected error processing Excel file: {str(e)}")

def generate_excel_template(student_names: list[str] | None = None) -> bytes:
    """Generate Excel template with proper column headers.

    If *student_names* is supplied the ФИО column is pre-filled with real
    student names and the score columns are left empty so the teacher only
    needs to fill in grades.  When no names are provided a small demo table
    with example values is produced instead.
    """
    if student_names:
        n = len(student_names)
        template_data = {
            'ФИО': student_names,
            'Процент за 1 предыдущий класс, %': [None] * n,
            'Q1, %': [None] * n,
            'Q2, %': [None] * n,
            'Q3, %': [None] * n,
            'Q4, %': [None] * n,
            'Учитель, %': [None] * n,
        }
    else:
        template_data = {
            'ФИО': ['Иванов Иван Иванович', 'Петров Петр Петрович', 'Сидорова Анна Владимировна'],
            'Процент за 1 предыдущий класс, %': [85.5, 92.0, 78.3],
            'Q1, %': [88, 90, 82],
            'Q2, %': [85, 94, 79],
            'Q3, %': [None, 89, 85],
            'Q4, %': [None, None, None],
            'Учитель, %': [87, 91, 80],
        }

    df = pd.DataFrame(template_data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Grades', index=False)

        worksheet = writer.sheets['Grades']

        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Light-blue fill for name column so it stands out
        name_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.fill = name_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = max(
                (len(str(cell.value)) for cell in column if cell.value is not None),
                default=8,
            )
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 3, 40)

    output.seek(0)
    return output.getvalue()



